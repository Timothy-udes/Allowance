import streamlit as st
import pandas as pd
import sqlite3
from io import BytesIO
from rapidfuzz import process, fuzz
import base64, os
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════
# PAGE CONFIG
# ═══════════════════════════════════════════════
st.set_page_config(
    page_title="Opti360 Driver Payment System",
    page_icon="💳",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# ═══════════════════════════════════════════════
# HIDE STREAMLIT DEFAULT MENU
# ═══════════════════════════════════════════════
hide_streamlit_style = """
    <style>
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header {visibility: hidden;}
    </style>
"""
st.markdown(hide_streamlit_style, unsafe_allow_html=True)

# ═══════════════════════════════════════════════
# LOGO HANDLER
# ═══════════════════════════════════════════════
def _img_to_b64(path):
    with open(path, "rb") as f:
        return base64.b64encode(f.read()).decode()

_HERE = os.path.dirname(os.path.abspath(__file__))
LOGO_OPTI360 = os.path.join(_HERE, "logo_opti360.png")
LOGO_CRISMEL = os.path.join(_HERE, "logo_crismel.png")

b64_opti = _img_to_b64(LOGO_OPTI360)
b64_crismel = _img_to_b64(LOGO_CRISMEL)

# ═══════════════════════════════════════════════
# UI DESIGN
# ═══════════════════════════════════════════════
st.markdown(f"""
<style>
body {{background:#F0F4F8;font-family:Inter;}}

.opti-header {{
    display:flex;justify-content:space-between;align-items:center;
    background:linear-gradient(135deg,#0D2137,#2471A3);
    padding:16px;border-radius:0 0 15px 15px;margin-bottom:20px;
}}
.opti-header img {{height:50px;}}
.opti-title {{color:white;font-size:1.5rem;font-weight:700;}}

.opti-card {{
    background:white;padding:20px;border-radius:12px;
    box-shadow:0 2px 10px rgba(0,0,0,0.08);
    margin-bottom:15px;
}}

.stat-pill {{
    background:#2471A3;color:white;padding:10px 18px;
    border-radius:10px;text-align:center;
}}

.stButton>button {{
    background:#1B3A5C;color:white;font-weight:700;
    border-radius:10px;padding:12px;
}}

[data-testid="stDownloadButton"]>button {{
    background:#1A5F3C;color:white;font-weight:700;
    border-radius:10px;padding:12px;
}}
</style>

<div class="opti-header">
    <div style="display:flex;align-items:center;gap:10px;">
        <img src="data:image/png;base64,{b64_opti}">
        <div class="opti-title">Driver Payment Automation System</div>
    </div>
    <div style="color:white;">💳 Professional Edition</div>
</div>
""", unsafe_allow_html=True)

# ═══════════════════════════════════════════════
# DATABASE
# ═══════════════════════════════════════════════
conn = sqlite3.connect("drivers.db")

# ═══════════════════════════════════════════════
# FILE READER
# ═══════════════════════════════════════════════
def read_file(file):
    if file.name.endswith(("xlsx","xls","xlsm")):
        return pd.read_excel(file)
    elif file.name.endswith("csv"):
        return pd.read_csv(file)
    elif file.name.endswith("ods"):
        return pd.read_excel(file, engine="odf")
    return None

# ═══════════════════════════════════════════════
# EXCEL FORMATTER
# ═══════════════════════════════════════════════
def to_excel(df):
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False)
        ws = writer.sheets["Sheet1"]

        for col in ws.columns:
            ws.column_dimensions[get_column_letter(col[0].column)].width = 20

        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1B3A5C")

    return output.getvalue()

# ═══════════════════════════════════════════════
# UI LAYOUT
# ═══════════════════════════════════════════════
col1, col2 = st.columns(2)

with col1:
    st.markdown('<div class="opti-card"><b>Upload Driver Database</b></div>', unsafe_allow_html=True)
    driver_file = st.file_uploader("Driver DB", type=["xlsx","csv","ods"])

with col2:
    st.markdown('<div class="opti-card"><b>Upload Driver Report</b></div>', unsafe_allow_html=True)
    report_file = st.file_uploader("Driver Report", type=["xlsx","csv","ods"])

# ═══════════════════════════════════════════════
# PROCESS DRIVER DATABASE
# ═══════════════════════════════════════════════
if driver_file:
    df = read_file(driver_file)
    df.columns = df.columns.str.upper().str.strip()

    # Automatically detect key columns
    if "FMS DRIVER'S NAME" in df.columns:
        df.rename(columns={"FMS DRIVER'S NAME":"DRIVER_NAME"}, inplace=True)
    if "ACCOUNT NAME" in df.columns:
        df.rename(columns={"ACCOUNT NAME":"ACCOUNT_NAME"}, inplace=True)
    if "ACCOUNT NO" in df.columns:
        df.rename(columns={"ACCOUNT NO":"ACCOUNT_NO"}, inplace=True)

    df["DRIVER_NAME"] = df["DRIVER_NAME"].astype(str).str.upper().str.strip()
    df["ACCOUNT_NO"] = df["ACCOUNT_NO"].astype(str).str.zfill(10)

    df.to_sql("drivers", conn, if_exists="replace", index=False)
    st.success("Driver Database Loaded")

# ═══════════════════════════════════════════════
# PROCESS REPORT
# ═══════════════════════════════════════════════
if report_file:
    df_r = read_file(report_file)
    df_r.columns = df_r.columns.str.upper().str.strip()

    if "DRIVER NAME" in df_r.columns:
        df_r.rename(columns={"DRIVER NAME":"DRIVER_NAME"}, inplace=True)
    if "TOTAL AMOUNT" in df_r.columns:
        df_r.rename(columns={"TOTAL AMOUNT":"AMOUNT"}, inplace=True)

    if "DRIVER_NAME" not in df_r.columns or "AMOUNT" not in df_r.columns:
        st.error("Required columns missing in the report: 'DRIVER NAME' and/or 'TOTAL AMOUNT'")
    else:
        df_r["DRIVER_NAME"] = df_r["DRIVER_NAME"].astype(str).str.upper().str.strip()
        df_r = df_r.groupby("DRIVER_NAME", as_index=False)["AMOUNT"].sum()

        df_db = pd.read_sql("SELECT * FROM drivers", conn)

        matches = []
        for name in df_r["DRIVER_NAME"]:
            match = process.extractOne(name, df_db["DRIVER_NAME"], scorer=fuzz.ratio)
            matches.append(match[0] if match and match[1] > 80 else None)

        df_r["MATCH"] = matches

        final = pd.merge(df_r, df_db, left_on="MATCH", right_on="DRIVER_NAME", how="left")

        final["ACCOUNT_NO"] = final["ACCOUNT_NO"].astype(str).str.zfill(10)
        final["S/N"] = range(1, len(final)+1)

        final = final[["S/N","DRIVER_NAME_x","AMOUNT","ACCOUNT_NAME","ACCOUNT_NO","BANK"]]
        final.columns = ["S/N","DRIVER NAME","AMOUNT","ACCOUNT NAME","ACCOUNT NO","BANK"]

        st.markdown('<div class="opti-card"><b>Final Payment Table</b></div>', unsafe_allow_html=True)
        st.dataframe(final)

        # Stats
        st.markdown(f"""
        <div style="display:flex;gap:10px;">
            <div class="stat-pill"><b>{len(final)}</b><br>Drivers</div>
            <div class="stat-pill"><b>₦{final['AMOUNT'].sum():,.0f}</b><br>Total Payment</div>
        </div>
        """, unsafe_allow_html=True)

        # Download
        st.download_button(
            "Download Payment Report",
            to_excel(final),
            file_name="Driver_Payments.xlsx"
        )

        bank = final[["ACCOUNT NAME","ACCOUNT NO","BANK","AMOUNT"]].dropna()

        st.download_button(
            "Download Bank Sheet",
            to_excel(bank),
            file_name="Bank_Payment.xlsx"
        )

# ═══════════════════════════════════════════════
# FOOTER
# ═══════════════════════════════════════════════
st.markdown(f"""
<div style="margin-top:40px;padding:20px;background:#1B3A5C;color:white;border-radius:10px;">
Opti360 Driver Payment System | Powered by Crismel Solutions
</div>
""", unsafe_allow_html=True)